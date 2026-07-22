import uuid
from datetime import date

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession

from app.db.session import get_db
from app.deps import get_current_user
from app.models.user import User
from app.schemas.report import ReportResponse
from app.services.report_service import generate_report

import io
import os
from openpyxl import Workbook

router = APIRouter(prefix="/api/reports", tags=["Reports"])


@router.get("", response_model=ReportResponse)
async def get_report(
    report_type: str = Query("daily", pattern="^(daily|weekly|monthly|yearly|custom)$"),
    start_date: date | None = Query(None),
    end_date: date | None = Query(None),
    bean_type_id: uuid.UUID | None = Query(None),
    db: AsyncSession = Depends(get_db),
    _user: User = Depends(get_current_user),
):
    """Generate a report based on type and date range."""
    try:
        data = await generate_report(db, report_type, start_date, end_date, bean_type_id)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    return ReportResponse(
        report_type=data["report_type"],
        start_date=date.fromisoformat(data["start_date"]),
        end_date=date.fromisoformat(data["end_date"]),
        data=data,
    )


@router.get("/export/excel")
async def export_report_excel(
    report_type: str = Query("daily", pattern="^(daily|weekly|monthly|yearly|custom)$"),
    start_date: date | None = Query(None),
    end_date: date | None = Query(None),
    bean_type_id: uuid.UUID | None = Query(None),
    db: AsyncSession = Depends(get_db),
    _user: User = Depends(get_current_user),
):
    """Export report as Excel file."""
    try:
        data = await generate_report(db, report_type, start_date, end_date, bean_type_id)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    # Header
    ws.append(["Beans Inventory Report"])
    ws.append([f"Type: {data['report_type']}"])
    ws.append([f"Period: {data['start_date']} to {data['end_date']}"])
    ws.append([])

    # Arrivals
    ws.append(["Arrivals Summary"])
    ws.append(["Count", "Bean Type", "Bags", "Total Cost"])
    for item in data["arrivals"]["bags_by_type"]:
        ws.append(["", item["bean_type_name"], item["quantity_bags"], ""])
    ws.append([data["arrivals"]["count"], "", "", data["arrivals"]["total_cost"]])
    ws.append([])

    # Sales
    ws.append(["Sales Summary"])
    ws.append(["Count", "Bean Type", "Bags", "Total Revenue"])
    for item in data["sales"]["bags_by_type"]:
        ws.append(["", item["bean_type_name"], item["quantity_bags"], ""])
    ws.append([data["sales"]["count"], "", "", data["sales"]["total_revenue"]])
    ws.append([])

    # Adjustments
    ws.append(["Adjustments Summary"])
    ws.append(["Count", "Bean Type", "Type", "Quantity (Viss)"])
    for item in data["adjustments"]["details"]:
        sign = "+" if item["adjustment_type"] == "increase" else "-"
        ws.append(["", item["bean_type_name"], sign, item["quantity_viss"]])
    ws.append([data["adjustments"]["count"], "", "", data["adjustments"]["total_quantity_viss"]])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=report_{report_type}.xlsx"},
    )


@router.get("/export/pdf")
async def export_report_pdf(
    report_type: str = Query("daily", pattern="^(daily|weekly|monthly|yearly|custom)$"),
    start_date: date | None = Query(None),
    end_date: date | None = Query(None),
    bean_type_id: uuid.UUID | None = Query(None),
    db: AsyncSession = Depends(get_db),
    _user: User = Depends(get_current_user),
):
    """Export report as PDF file with Myanmar font support."""
    try:
        data = await generate_report(db, report_type, start_date, end_date, bean_type_id)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    font_path = os.path.join(
        os.path.dirname(__file__), "..", "..", "fonts", "NotoSansMyanmar-Regular.ttf"
    )

    from fpdf import FPDF

    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=20)

    # Register Myanmar font
    pdf.add_font("Myanmar", "", font_path)

    # ── Helper: write a cell using Helvetica for Latin, Myanmar font for Myanmar text ──
    def write_mm_text(pdf, text: str, size: float = 10, style: str = "", align: str = "L"):
        """Write a line of text that may contain Myanmar characters."""
        pdf.set_font("Helvetica", style, size)
        w = pdf.get_string_width(text)
        pdf.cell(w, size * 0.35, text, new_x="LMARGIN", new_y="NEXT", align=align)

    def mm_cell(pdf, w, h, text: str, align: str = "L", border: int = 0):
        """Cell that switches font per character to handle Myanmar + Latin mixed text."""
        pdf.set_font("Myanmar", "", h)
        pdf.cell(w, h * 0.35, text, align=align, border=border)

    # ── Page header ──
    pdf.add_page()

    # Title - Helvetica
    pdf.set_font("Helvetica", "B", 18)
    pdf.cell(0, 12, "Beans Inventory Report", align="C", new_x="LMARGIN", new_y="NEXT")

    # Subtitle
    pdf.set_font("Helvetica", "", 9)
    period_text = f"Type: {data['report_type'].upper()}  |  Period: {data['start_date']} to {data['end_date']}"
    pdf.cell(0, 6, period_text, align="C", new_x="LMARGIN", new_y="NEXT")

    pdf.ln(4)

    # ── Arrivals Summary ──
    pdf.set_font("Helvetica", "B", 13)
    pdf.set_fill_color(220, 240, 220)
    pdf.cell(0, 8, "  Arrivals Summary", new_x="LMARGIN", new_y="NEXT", fill=True)
    pdf.ln(2)

    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 6, f"Count: {data['arrivals']['count']}", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(1)

    # Arrivals table header
    col_w = [90, 30, 30]
    pdf.set_font("Helvetica", "B", 9)
    pdf.set_fill_color(245, 245, 245)
    pdf.cell(col_w[0], 6, "  Bean Type", border=1, fill=True)
    pdf.cell(col_w[1], 6, "Bags", border=1, align="C", fill=True)
    pdf.cell(col_w[2], 6, "Amount", border=1, align="R", fill=True)
    pdf.ln()

    # Arrivals data rows
    pdf.set_font("Myanmar", "", 9)
    for item in data["arrivals"]["bags_by_type"]:
        name = item["bean_type_name"]
        pdf.cell(col_w[0], 6, "  " + name, border=1)
        pdf.set_font("Helvetica", "", 9)
        pdf.cell(col_w[1], 6, f"{item['quantity_bags']:,}", border=1, align="C")
        pdf.cell(col_w[2], 6, "", border=1, align="R")
        pdf.ln()
        pdf.set_font("Myanmar", "", 9)

    # Arrivals total
    pdf.set_font("Helvetica", "B", 9)
    pdf.cell(col_w[0] + col_w[1], 6, "  Total Cost", border="TB", align="L")
    pdf.cell(col_w[2], 6, f"{data['arrivals']['total_cost']:,.0f}", border="TB", align="R")
    pdf.ln(8)

    # ── Sales Summary ──
    pdf.set_font("Helvetica", "B", 13)
    pdf.set_fill_color(220, 230, 245)
    pdf.cell(0, 8, "  Sales Summary", new_x="LMARGIN", new_y="NEXT", fill=True)
    pdf.ln(2)

    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 6, f"Count: {data['sales']['count']}", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(1)

    # Sales table header
    pdf.set_font("Helvetica", "B", 9)
    pdf.set_fill_color(245, 245, 245)
    pdf.cell(col_w[0], 6, "  Bean Type", border=1, fill=True)
    pdf.cell(col_w[1], 6, "Bags", border=1, align="C", fill=True)
    pdf.cell(col_w[2], 6, "Revenue", border=1, align="R", fill=True)
    pdf.ln()

    # Sales data rows
    pdf.set_font("Myanmar", "", 9)
    for item in data["sales"]["bags_by_type"]:
        name = item["bean_type_name"]
        pdf.cell(col_w[0], 6, "  " + name, border=1)
        pdf.set_font("Helvetica", "", 9)
        pdf.cell(col_w[1], 6, f"{item['quantity_bags']:,}", border=1, align="C")
        pdf.cell(col_w[2], 6, "", border=1, align="R")
        pdf.ln()
        pdf.set_font("Myanmar", "", 9)

    # Sales total
    pdf.set_font("Helvetica", "B", 9)
    pdf.cell(col_w[0] + col_w[1], 6, "  Total Revenue", border="TB", align="L")
    pdf.cell(col_w[2], 6, f"{data['sales']['total_revenue']:,.0f}", border="TB", align="R")
    pdf.ln(8)

    # ── Adjustments Summary ──
    pdf.set_font("Helvetica", "B", 13)
    pdf.set_fill_color(255, 245, 220)
    pdf.cell(0, 8, "  Adjustments Summary", new_x="LMARGIN", new_y="NEXT", fill=True)
    pdf.ln(2)

    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 6, f"Count: {data['adjustments']['count']}", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(1)

    # Adjustments table header
    cols_adj = [90, 30, 30]
    pdf.set_font("Helvetica", "B", 9)
    pdf.set_fill_color(245, 245, 245)
    pdf.cell(cols_adj[0], 6, "  Bean Type", border=1, fill=True)
    pdf.cell(cols_adj[1], 6, "Type", border=1, align="C", fill=True)
    pdf.cell(cols_adj[2], 6, "Viss", border=1, align="R", fill=True)
    pdf.ln()

    # Adjustments data rows
    pdf.set_font("Myanmar", "", 9)
    for item in data["adjustments"]["details"]:
        name = item["bean_type_name"]
        sign = "+" if item["adjustment_type"] == "increase" else "-"
        adj_type = "Increase" if item["adjustment_type"] == "increase" else "Decrease"
        pdf.cell(cols_adj[0], 6, "  " + name, border=1)
        pdf.set_font("Helvetica", "", 9)
        pdf.cell(cols_adj[1], 6, adj_type, border=1, align="C")
        pdf.cell(cols_adj[2], 6, f"{sign}{item['quantity_viss']:,.2f}", border=1, align="R")
        pdf.ln()
        pdf.set_font("Myanmar", "", 9)

    # Adjustments total
    pdf.set_font("Helvetica", "B", 9)
    pdf.cell(cols_adj[0] + cols_adj[1], 6, "  Total", border="TB", align="L")
    pdf.cell(
        cols_adj[2], 6, f"{data['adjustments']['total_quantity_viss']:,.2f} Viss",
        border="TB", align="R",
    )
    pdf.ln(8)

    # ── Footer ──
    pdf.set_font("Helvetica", "", 7)
    pdf.set_text_color(150, 150, 150)
    pdf.cell(0, 4, f"Generated on {date.today().isoformat()} | Beans Inventory Management System", align="C")

    output = io.BytesIO()
    pdf.output(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=report_{report_type}.pdf"},
    )
